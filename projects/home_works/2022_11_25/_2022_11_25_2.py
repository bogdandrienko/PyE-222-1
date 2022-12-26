# 2)	Записать данные в другой файл, но уже вертикально, объяснение на картинках:

import openpyxl


def writing(matrix: list[list[any]], new_filename="new.xlsx", is_logging=False) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # worksheet["A1"] = "Марина"
    # worksheet.cell(2, 2, "Алема")
    # rows = [            # TODO Rows
    #     [1, 2, 3],      # TODO Row 1
    #     ["A", "B", "C"] # TODO Row 2
    # ]

    # A 1
    # B 2
    # C 3

    # 1 1
    # 2 1
    # 1 3

    for index_row, row in enumerate(matrix, 1):
        for index_column, cell in enumerate(row, 1):
            worksheet.cell(index_column, index_row, cell)

    workbook.save(new_filename)
    workbook.close()
