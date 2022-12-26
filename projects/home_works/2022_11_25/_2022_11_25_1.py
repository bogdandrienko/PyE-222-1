# 1)	Прочитать файл в матрицу(двумерный массив)

import openpyxl
from openpyxl.utils import get_column_letter


def reading(filename: str, is_logging=False) -> list[list[any]]:
    workbook = openpyxl.load_workbook(filename=filename)
    # worksheet = workbook["list1"]  # TODO activate by sheet name
    worksheet = workbook.active

    # a1 = worksheet["A1"].value
    # b2 = worksheet.cell(2, 2).value
    # print(a1, b2)

    row1 = []
    for i in range(1, worksheet.max_column + 1):
        char = get_column_letter(i)
        val1 = worksheet[f"{char}1"].value
        row1.append(val1)

    row2 = []
    for i in range(1, worksheet.max_column + 1):
        char = get_column_letter(i)
        val2 = worksheet[f"{char}2"].value
        row2.append(val2)

    # rows = []
    # for i in range(1, worksheet.max_row + 1):
    #     row = []
    #     for j in range(1, worksheet.max_column + 1):
    #         val = worksheet.cell(i, j).value
    #         row.append(val)
    #     rows.append(row)

    if is_logging:
        print(row1)
        print(row2)

    matrix = [row1, row2]
    workbook.close()
    return matrix


if __name__ == '__main__':
    reading("Лист Microsoft Excel.xlsx")
