import openpyxl
# исходные данные
files = ["Лист1.xlsx", "Лист2.xlsx", "Лист3.xlsx"]
# читаем файлы в матрицу
matrix = []
for index_col, file in enumerate(files, 1):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active
    column = []
    for row_index in range(1, worksheet.max_row+1):
        column.append(worksheet.cell(row_index, index_col).value)
    matrix.append(column)
    workbook.close()
# записываем матрицу в файл
workbook = openpyxl.Workbook()
worksheet = workbook.active
for index_r, row in enumerate(matrix, 1):
    for index_c, cell in enumerate(row, 1):
        worksheet.cell(index_c, index_r, cell)
workbook.save("new.xlsx")