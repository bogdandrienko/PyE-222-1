import datetime
import re
import time

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import sqlite3
import hashlib
import openpyxl
from openpyxl.styles import PatternFill


def home(request):
    return render(request, "home.html")


# todo CRUD
# C - create    - INSERT
# R - read      - SELECT
# U - update    - UPDATE / UPSERT
# D - delete    - DELETE


"""

"""


def api_users_import(request):
    if request.method == "GET":

        # получить файл от пользователя
        # конвертировать в массив словарей
        # создать в базе данных пользователей из массива

        filename = "static/export_users/пользователи.xlsx"
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active

        matrix = []
        for row in range(2, worksheet.max_row + 1):
            sex = str(worksheet[f'E{row}'].value)
            if sex == "Муж.":
                sex = "Мужской"
            elif sex == "Жен.":
                sex = "Женский"
            else:
                sex = "Скрыто"
            new_people = {
                "username": str(worksheet[f'B{row}'].value),
                "password": str(worksheet[f'C{row}'].value),
                "email": str(worksheet[f'D{row}'].value),
                "sex": sex,
            }
            matrix.append(new_people)
        print(matrix)

        # + 1 connection = 0.1
        # + 1 cursor = 0.075
        # + 1 user = 0.05

        # ! 1 000 000 * user - МНОГО

        # каждый раз открывать (connection) базу данных
        # один раз открыть, но каждый раз делать отдельный запрос (cursor)
        # один раз открыть, отправить один запрос (cursor)
        # один раз открыть, и в зависимости от количества пользователей "пачками" отправлять в одном запросе

        # todo VERY BAD - каждый раз открывать (connection) базу данных для каждого пользователя
        # for user in matrix:
        #     with sqlite3.connect('database/db.db') as connection:
        #         cursor = connection.cursor()
        #         query = """
        #     INSERT INTO users (username, email, password, sex) VALUES (?, ?, ?, ?)
        #                 """
        #         hash_password = hashlib.sha256(user['password'].encode()).hexdigest()
        #         args = (user['username'], user['email'], hash_password, user['sex'])
        #         cursor.execute(query, args)
        #         connection.commit()

        # todo CONTEXT - bad - один раз открыть, но каждый раз делать отдельный запрос (cursor)
        # with sqlite3.connect('database/db.db') as connection:
        #     cursor = connection.cursor()
        #     for user in matrix:
        #         query = """
        #     INSERT INTO users (username, email, password, sex) VALUES (?, ?, ?, ?)
        #                 """
        #         hash_password = hashlib.sha256(user['password'].encode()).hexdigest()
        #         args = (user['username'], user['email'], hash_password, user['sex'])
        #         cursor.execute(query, args)
        #         connection.commit()

        # todo CONTEXT - good - один раз открыть, и в зависимости от количества пользователей "пачками" отправлять в одном запросе
        with sqlite3.connect('database/db.db') as connection:
            cursor = connection.cursor()

            matrix = ["user1", "user2", "user3"]
            matrix = [
                ["user1", "user2"],     # group 1
                # time.sleep
                ["user3"],              # group 2
                # time.sleep
                ["user4", "user5"]      # group 3
            ]

            for user in matrix:
                query = """
            INSERT INTO users (username, email, password, sex) VALUES (?, ?, ?, ?)
                        """
                hash_password = hashlib.sha256(user['password'].encode()).hexdigest()
                args = (user['username'], user['email'], hash_password, user['sex'])
                cursor.execute(query, args)
                connection.commit()

                time.sleep(1.0)

        return HttpResponse("Метод не реализован")


def api_users_export(request):
    if request.method == "GET":
        with sqlite3.connect('database/db.db') as connection:
            cursor = connection.cursor()
            cursor.execute("""SELECT id, username, email from users order by id DESC""")
            rows = cursor.fetchall()
        # File -> Workbook -> Worksheet's -> Row's | Column's -> Cell
        workbook = openpyxl.load_workbook("static/export_users/отчёт.xlsx")  # todo NEW FILE | load_workbook - чтение существующего
        worksheet = workbook["Сотрудники"]
        for row_idx, row in enumerate(rows, 2):
            worksheet[f'A{row_idx}'] = row[0]
            worksheet[f'B{row_idx}'] = row[1]
            worksheet[f'C{row_idx}'] = row[2]
        # worksheet["A1"].fill = PatternFill(fill_type='solid', start_color='0070c1', end_color='0070c1', fgColor='0070c1', bgColor='0070c1')
        # worksheet["A2"].fill = PatternFill("solid", fgColor="DDDDDD")
        worksheet["C1"].font = openpyxl.styles.Font(color='0070c1')
        # STATIC_ROOT = '/home/andriyenko1bogdan/django_project/static'
        filename = f"export_users/temp/отчёт_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        workbook.save(f"static/{filename}")
        return render(request, "export.html", context={"filename": filename})


def register(request):
    if request.method == "GET":
        return render(request, "register.html")
    elif request.method == "POST":
        # print(request.POST)
        # TODO НУЖНО ВСЕГДА ПРОВЕРЯТЬ ДАННЫЕ

        email = request.POST.get('email')
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
            raise Exception(f"Ошибка ввода почты {email}")

        password = request.POST.get('password')
        if not re.match(r"^(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{12,}$", password):
            raise Exception(f"Ошибка ввода пароля {password}")

        username = request.POST.get('username')
        save = request.POST.get('save')
        sex = request.POST.get('sex')
        datetime = request.POST.get('datetime')  # .strftime("%Y-%m-%d %H:%M:%S")

        # todo WRITE TO TXT FILE
        # with open("static/users.txt", mode="a", encoding="utf-8") as file:
        #     file.write(f"{email} | {password} | {username} | {save} | {sex} | {datetime} |\n")

        with sqlite3.connect('database/db.db') as connection:
            cursor = connection.cursor()

            query = """
INSERT INTO users (username, email, password, sex) VALUES (?, ?, ?, ?)
            """
            # todo ОТКРЫТЫЕ ПАРОЛИ В БАЗЕ ХРАНИТЬ НЕЛЬЗЯ!
            hash_password = hashlib.sha256(password.encode()).hexdigest()
            args = (username, email, hash_password, sex)

            cursor.execute(query, args)
            connection.commit()
        return HttpResponse("Метод в процессе реализации")
    else:
        return HttpResponse("Метод не реализован")
