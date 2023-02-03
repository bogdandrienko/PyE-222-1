# pip install "fastapi[all]"
# pip install "uvicorn[standard]"
# pip install psycopg2
import json

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import psycopg2
import openpyxl

from starlette.responses import RedirectResponse

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ipconfig (192.168.0.128)
# uvicorn main:app --reload --host=0.0.0.0 --port=8000
# http://127.0.0.1:8000/

@app.get("/get_string")
async def get_string():
    return "get_string"


@app.get("/get_json")
async def get_json():
    return {"detail": "get_json"}


@app.get("/get_html", response_class=HTMLResponse)
async def get_html(request: Request):
    return templates.TemplateResponse("Home.html", {"request": request})


class Product:
    def __init__(self, id_: int, title: str, price: float, count: int, type_measure: str, nomeklatura_id: str):
        self.id_ = id_
        self.title = title
        self.price = price
        self.count = count
        self.type_measure = type_measure
        self.nomeklatura_id = nomeklatura_id

    @staticmethod
    def create_product_from_row_from_db(row: tuple):
        return Product(id_=row[0], title=row[1], price=row[2], count=row[3], type_measure=row[4], nomeklatura_id=row[5])

    # @staticmethod
    # def ph_create_connection_to_products_db(query: str):
    #     with psycopg2.connect(user="pgs_usr", password="12345Qwerty!",
    #                           host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
    #         with connection.cursor() as cursor:
    #             # SELECT * FROM products;
    #             return cursor

    @staticmethod
    def select_rows_query(query="select * from products;") -> list[tuple]:
        with psycopg2.connect(user="pgs_usr", password="12345Qwerty!",
                              host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
            with connection.cursor() as cursor:
                cursor.execute(query)
                rows = cursor.fetchall()
                if rows is None:
                    raise Exception("not have objs")
                return rows

    @staticmethod
    def select_row(query="select * from product where id=1;") -> tuple:
        with psycopg2.connect(user="pgs_usr", password="12345Qwerty!",
                              host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
            with connection.cursor() as cursor:
                cursor.execute(query)
                row = cursor.fetchone()
                if row is None:
                    raise Exception("not have objs")
                return row

    @staticmethod
    def insert_new_product(title: str, price: float, count: float, type_measure: str, nomeklatura_id: str) -> bool:
        # todo ВАЛИДАЦИЯ
        if len(title) < 3:
            raise Exception("Слишком")
        with psycopg2.connect(user="pgs_usr", password="12345Qwerty!",
                              host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
            connection.autocommit = False
            with connection.cursor() as cursor:
                status = False
                try:
                    cursor.execute("insert into products(title, price, count, type_measure, nomeklatura_id) VALUES"
                                   f"('{title}', '{price}', '{count}', '{type_measure}', '{nomeklatura_id}');")
                except Exception as error:
                    print("error", error)
                    connection.rollback()
                else:
                    connection.commit()
                    status = True
                finally:
                    return status

    @staticmethod
    def delete_product(title: str) -> bool:
        # todo ВАЛИДАЦИЯ
        if len(title) < 3:
            raise Exception("Слишком")
        with psycopg2.connect(user="pgs_usr", password="12345Qwerty!",
                              host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
            connection.autocommit = False
            with connection.cursor() as cursor:
                status = False
                try:
                    cursor.execute(f"delete from products where title='{title}';")
                except Exception as error:
                    print("error", error)
                    connection.rollback()
                else:
                    connection.commit()
                    status = True
                finally:
                    return status

    @staticmethod
    def get_all_product_types() -> list[tuple]:
        return [("l", "литры"), ("kg", "килограммы"), ("sh", "штуки"), ("u", "упаковки")]

    def get_dict(self) -> dict:
        return {"title": self.title, "price": self.price, "count": self.count}

    def get_txt(self) -> str:
        return f"title: {self.title} | price: {self.price}"

    def save_obj_in_txt(self):
        with open(f"{self.title}.txt", "w", encoding="utf-8") as file:
            file.write(self.get_txt())

    def serialize_one(self):
        with open(f"{self.title}.json", "w", encoding="utf-8") as file:
            json.dump(self.get_dict(), file)

    def save_one_to_excel(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.cell(row=1, column=1, value="Наименование")
        worksheet.cell(row=1, column=2, value="Цена")
        worksheet.cell(row=1, column=3, value="Количество")

        worksheet.cell(row=2, column=1, value=self.title)
        worksheet.cell(row=2, column=2, value=self.price)
        worksheet.cell(row=2, column=3, value=self.count)

        workbook.save(f"{self.title}.xlsx")


class Products:
    def __init__(self, products: list[Product]):
        self.products = products

    def serialize_many(self) -> None:
        objs = [x.get_dict() for x in self.products]
        with open(f"backup.json", "w", encoding="utf-8") as file:
            json.dump(objs, file)

    def save_many_to_excel(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet.cell(row=1, column=1, value="Наименование")
        worksheet.cell(row=1, column=2, value="Цена")
        worksheet.cell(row=1, column=3, value="Количество")

        for indx, i in enumerate(self.products, 2):
            worksheet.cell(row=indx, column=1, value=i.title)
            worksheet.cell(row=indx, column=2, value=i.price)
            worksheet.cell(row=indx, column=3, value=i.count)

        workbook.save(f"backup.xlsx")


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    rows = Product.select_rows_query(
        "SELECT title, price, count, type_measure FROM products order by count desc, title asc;"
    )
    products = [
        {
            "title": x[0],
            "price": x[1],
            "count": x[2],
            "type_measure": x[3],
        }
        for x in rows
    ]
    context = {"request": request, "products": products, "types": Product.get_all_product_types()}
    return templates.TemplateResponse("ProductAll.html", context)


@app.post("/create")
async def create_post(request: Request):
    form = await request.form()

    title = form.get("title")
    count = form.get("count")
    price = form.get("price")
    type_measure = form.get("type_measure")
    nomeklatura_id = title + "_" + type_measure
    Product.insert_new_product(
        title=title, price=price, count=count, type_measure=type_measure, nomeklatura_id=nomeklatura_id
    )
    return RedirectResponse("/", status_code=303)


@app.get("/delete/{title}")
async def delete_post(request: Request, title: str):
    result = Product.delete_product(title=title)
    if result:
        return "Продукт успешно удалён"
    return f"Продукт не удалён"


@app.get("/backup")
async def backup(request: Request):
    result = Product.select_rows_query()
    # product_class_list = []
    # for i in result:
    #     if len(i[1]) > 3:
    #         new_product = Product.create_product_from_row_from_db(row=i)
    #         product_class_list.append(new_product)
    product_class_list = [Product.create_product_from_row_from_db(row=i) for i in result if len(i[1]) > 3]
    for i in product_class_list:
        # i.save_obj_in_txt()
        i.save_one_to_excel()
    prds = Products(product_class_list)
    prds.serialize_many()
    prds.save_many_to_excel()
    return f""


def database():
    # cd C:\Program Files\PostgreSQL\15\bin>
    # cmd
    # psql -U postgres
    # \l
    # \d
    # CREATE USER pgs_usr WITH PASSWORD '12345Qwerty!';
    # CREATE DATABASE pgs_db OWNER pgs_usr;
    # \connect pgs_db
    # CREATE TABLE public.products ( id serial PRIMARY KEY, title VARCHAR(128) unique NOT NULL, price double precision DEFAULT 0.0, count INT default 0, type_measure VARCHAR(10) DEFAULT 'kg', nomeklatura_id VARCHAR(255) );
    # \d
    # GRANT ALL PRIVILEGES ON DATABASE pgs_db TO pgs_usr;
    # GRANT ALL PRIVILEGES ON ALL TABLES IN SCHEMA public to pgs_usr;
    # GRANT ALL PRIVILEGES ON ALL SEQUENCES IN SCHEMA public to pgs_usr;
    # GRANT ALL PRIVILEGES ON ALL FUNCTIONS IN SCHEMA public to pgs_usr;

    # select * from products;
    # insert into products (title, price, count, type_measure, nomeklatura_id) VALUES ('Бананы', '1200.07', '60', 'kg', '3_Бананы');
    # insert into products (title, price, count, type_measure, nomeklatura_id) VALUES ('Olives', '3600.00', '7', 'kg', '5_Olive');
    # insert into products (title, price, count, type_measure, nomeklatura_id) VALUES ('Ananas', '6000.00', '0', 'kg', '666_Ananas');
    # select * from products;

    # delete from products where id=1;

    # \q
    pass
