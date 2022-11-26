from math import ceil, floor
from typing import Union

from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

# xlsx xls xlsm

# 1) положим их в одну папку и переименуем на латинницу
# 2) найти как прочитать их начинку (две переменные с данными)   Read
# 3) две переменные => одну переменную
# 4) создать новый документ с новыми данными                     Create
# 5) обновить документ с новыми данными                          Update


# CRUD
# Create - документа ещё нет, нужно его создать
# Read - документ уже есть, нужно прочитать
# Update - документ уже есть, нужно изменить данные
# Delete


workbook = openpyxl.load_workbook("data/data1.xlsx")
worksheet = workbook.active  # workbook["Лист 2"]

# dict1 = {"name": "Python"}
# dict1["name"]

# cella1 = worksheet["B10"].value
# print(cella1, type(cella1))
#
# cella2 = worksheet.cell(10, 2).value
# print(cella2, type(cella2))

arr1 = []
#           start, stop, step
#               range(20) # [0....19]
# for i in [2, 3, 4, 5, 1]:  #
for i in range(2, 14 + 1, 1):  # [2, 3, 4, ... 14]
    cell_value = worksheet[f"B{i}"].value
    arr1.append(cell_value)
    #
#
# print(arr1, type(arr1))

arr2 = []
# for char in ["A", "B", "C"]:  # [2, 3, 4, ... 14]
# for char in "ABCDE":  # [2, 3, 4, ... 14]
# for char in range(1, 4+1, 1):  # [2, 3, 4, ... 14]
#     # cell_value = worksheet.cell(2, char).value
#     val1 = get_column_letter(char)
#     print(val1)
#
#     cell_value = worksheet[f"{val1}2"].value
#     # cell_value = worksheet[f"{char}2"].value
#     arr2.append(cell_value)
#     #
# #
# print(arr2, type(arr2))

# max_row = worksheet.max_row
# max_column = worksheet.max_column


# arr5 = [1, 2, 3, 5]
# arr6 = [[1, [1, 2]], [1, 2], [1, 2], [1, 2]]
# todo создаём внешний пустой массив
external_array1 = []

# todo проходим циклом по строкам
for row in range(1, worksheet.max_row + 1):

    # todo создаём внутренний пустой массив
    internal_array = []

    # todo проходим циклом по столбцам
    for column in range(1, worksheet.max_column + 1):
        # todo наполняем внутренний массив значения ячеек
        internal_array.append(worksheet.cell(row, column).value)

    # todo наполняем внешний массив массивом со значением строк
    external_array1.append(internal_array)

# print(external_array1)
# for i in external_array1:
#     print(i)

workbook2 = openpyxl.load_workbook("data/data2.xlsx")
worksheet2 = workbook2.active

# todo создаём внешний пустой массив
external_array2 = []

# todo проходим циклом по столбцам
for column in range(1, worksheet2.max_column + 1):

    # todo создаём внутренний пустой массив
    internal_array2 = []

    # todo проходим циклом по строкам
    for row in range(2, worksheet2.max_row + 1):
        # todo наполняем внутренний массив значения ячеек
        internal_array2.append(worksheet2.cell(row, column).value)

    # todo наполняем внешний массив массивом со значением строк
    external_array2.append(internal_array2)

# print(external_array2)
# for i in external_array2:
#     print(i)

external_array1.extend(external_array2)
# print(external_array1)
# for i in external_array1:
#     print(i)

# todo создание новой рабочей книги в оперативной памяти
workbook3 = Workbook()

# todo выбор активного рабочего листа
worksheet3 = workbook3.active

# todo присвоение
# worksheet3['B3'] = 42
# worksheet3.cell(5, 3, 666)
#
# for column in range(1, 1000):
#     worksheet3.cell(3, column, 1000-column)
#
# for row in range(1, 300):
#     worksheet3.cell(row, 2, "Python")
#
# name = "Alema"
# for row in range(8, 12+1):
#     for column in range(4, 7+1):
#         worksheet3.cell(row, column, name)

print(external_array1)

external_counter = 0
# todo проходим по внешнему массиву, каждый раз берём внутренний массив
for external in external_array1:  # ['И.И.И.', 30, 600000]
    external_counter += 1

    internal_counter = 0
    # todo проходим по внутреннему массиву, каждый раз берём значение
    for internal in external:  # 'И.И.И.'  30  600000
        internal_counter += 1

        # print(f"строка: {external_counter}, колонка: {internal_counter}, значение: {internal}")

        # todo устанавливаем значени в ячейку
        worksheet3.cell(external_counter, internal_counter, internal)
#      0   1   2  3  4  !5  !6
arr = [5, "P", 2, 9, 6, 5, "P"]
for i in arr:
    index = arr.index(i)
    # print(i, index)
i_counter = 0
for i in arr:
    # print(i, i_counter)

    i_counter += 1  # increment
    #
#

# unpacking - распаковка значений
print(arr)
for index, value in enumerate(arr, 666):  # enumerate - получает значение, а возвращает индекс и значение
    print(index, value)

# [5, 'P', 2, 9, 6, 5, 'P']
# [(0, 5), (1, 'P'), (2, 2)...]

# todo сохранение
workbook3.save("data/new_data.xlsx")


def double(val: Union[int, float]) -> Union[int]:
    return round(val + 2)


res = double(2.0)
print(res)

arr9 = [1, 2, 3, 4, 5, 6]
# for i in arr9:
#     result = double(val=i)
#     print(result)

# val3 = map(double, arr9)  # lazy computing - только тогда, когда его вызывают
# print(list(val3))

for i in map(double, arr9):
    print(i)


# d = float(input("Введи диаметр"))
# l = 3.14 * d
# print(l)

r = 2*3.14*5 / 2*3.14

print(r)
print(floor(r))
