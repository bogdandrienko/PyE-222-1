# 3) Записывать в разные excel файлы
import openpyxl


def third(final_data: list[list[any]], is_logging=False) -> None:
    if is_logging:
        print(final_data)
    example_many_excels = \
        [
            [
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed11111111111111"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
            ],
            [
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed2222222222222"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
            ],
            [
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed33333333333"],
                ["userId", "id", "title", "completed"],
                ["userId", "id", "title", "completed"],
            ],
        ]

    example_one_excels = example_many_excels[2]

    def example_one():
        _new_workbook = openpyxl.Workbook()
        _new_worksheet = new_workbook.active

        _index_row = 0
        for _row in example_one_excels:
            _index_row = _index_row + 1
            print(_row)
            for _index_column, _cell in enumerate(_row, 1):
                print(_cell)
                new_worksheet.cell(row=_index_row, column=_index_column, value=_cell)

        _new_workbook.save(f"data/new0.xlsx")

    # example_one()

    for excel_index, excel in enumerate(final_data, 1):
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        index_row = 0
        for row in excel:
            index_row = index_row + 1
            for index_column, cell in enumerate(row, 1):
                new_worksheet.cell(row=index_row, column=index_column, value=cell)
        #                                200 199 192
        new_workbook.save(f"data/new{201-excel_index}.xlsx")


if __name__ == "__main__":  # TODO вызывается только если файл является основным
    third(final_data=[[""]])
