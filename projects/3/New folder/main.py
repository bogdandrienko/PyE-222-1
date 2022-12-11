import openpyxl
from openpyxl.utils import get_column_letter
import datetime

# print("Hello World!")

# 1) (Гульнура) Получить данные с текстовых документов (одна папка, но названия рандомные) => сырая матрица
# 2) (Марина) сырая матрица [["24", "Б"...], ["23", "и"]] => в матрицу финальную
# 3) (Диас) [["Б", "24"...], ["23", "и"]] Записать в шаблон (уже существующий эксель файл)

# file = open("data/data2.txt", "r", encoding="utf8")
# text1 = file.readlines()
# print(text1)
# file.close()

import os

files = []
for dirpath, dirnames, filenames in os.walk("data"):
    # print(filenames)
    files = filenames
    # for name in filenames:
    #     print(name)

# print(files)

# list1 = [1, 2, 3]
# list2 = [4, 5, 6]
# list1.extend(list2)
# print(list1)

lines = []
for file in files:
    with open(f"data/{file}", "r", encoding="utf8") as opened_file:
        text1 = opened_file.readlines()
        lines.extend(text1)

print(lines)


a = 5
b = 6

temp = a
a = b
b = temp

a, b = b, a
a1, b1 = "1", "2"

rows = []
for line in lines:
    print(line, len(line))
    line = line.strip().replace(" ", "")
    if len(line) <= 1:
        continue
    row = line.split(sep=",")
    print(row)
    row[0], row[1], row[2], row[3] = row[1], row[2], int(row[0]), int(row[3])
    print(row)
    rows.append(row)

print(rows)
for i in rows:
    print(i)

filename = "data.xlsx"
workbook = openpyxl.load_workbook(filename)
worksheet = workbook.active

for row_index, row in enumerate(rows, 1+1):
    for column_index, value in enumerate(row, 1):
        # print(f"{row_index}  |  {column_index}  |  {value}")
        # worksheet[f"{get_column_letter(column_index)}{row_index}"] = value

        worksheet.cell(row_index, column_index, value)


workbook.save(f"new_{datetime.datetime.now().strftime('%d-%m %H-%M')}_{filename}")
