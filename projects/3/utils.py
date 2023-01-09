import datetime
from concurrent.futures import ThreadPoolExecutor

import openpyxl
import threading


class Main:
    def exam(self):
        pass

    class Datetime:
        @staticmethod
        def get_current_time():
            return datetime.datetime.now()

        @staticmethod
        def get_asia_time():
            return datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")

        @staticmethod
        def get_time_in_selected_timezone(grinvich=6.0, cite_name="Asia/Almaty"):
            match cite_name:
                case "Asia/Almaty":
                    return datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=6)
                case "Asia/Oral":
                    return datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=5)
                case "Asia/Yekaterinburg":
                    return datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=5)
            # dt.astimezone()
            # 00:00 - япония (+9)
            # 21:00 - Астана (+6)
            # 18:00 - Турция (+3)
            # 18:00 - Москва (+3)
            # datetime.datetime.now() - локальное время компьютера
            return datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=grinvich)

        @staticmethod
        def example1():
            """Пример получения даты и времени для Китая"""

            print(Main.Datetime.get_time_in_selected_timezone(grinvich=8))

    class Excel:
        def __init__(self, filename=None, sheet_name=""):
            if filename is not None:
                self.workbook = openpyxl.load_workbook(filename)
            else:
                self.workbook = openpyxl.Workbook()
            if sheet_name:
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
            self.filename = filename
            self.sheet_name = sheet_name
            self.max_row = self.worksheet.max_row
            self.max_column = self.worksheet.max_column

        def read_all(self) -> list[list[any]]:
            rows = []
            for selected_row in range(1, self.max_row + 1):
                row = []
                for selected_column in range(1, self.max_column + 1):
                    value = self.worksheet.cell(selected_row, selected_column).value
                    if value is None:
                        row.append("")
                    else:
                        row.append(value)
                rows.append(row)
            return rows

        def read_selected_row(self, selected_row: int) -> list[any]:
            row = []
            for j in range(1, self.max_column + 1):
                value = self.worksheet.cell(selected_row, j).value
                if value is None:
                    row.append("")
                else:
                    row.append(value)
            return row

        def write_from_coordinates(self, matrix: list[list[any]], start_row=1, start_column=1) -> bool:
            try:
                rows = matrix
                for selected_row, row in enumerate(rows, start_row):
                    for selected_column, cell in enumerate(row, start_column):
                        self.worksheet.cell(selected_row, selected_column, cell)
                self.workbook.save(filename=self.filename)
                return True
            except Exception as error:
                print(error)
                return False

    class MyMultithreading:
        @staticmethod
        def start_in_side_threads(link_to_function, objects, max_workers=8):
            # 20
            # 1 vs 1 s
            # 20 / 1 = 20 s
            # 20 / 5 = 4 s
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                for idx, obj in enumerate(objects, 1):
                    executor.submit(link_to_function, f"{idx}.txt", obj)

            # TODO ресурсоёмко при более чем 50 задачах!!!
            # thread_task_list = []
            # for idx, obj in enumerate(objects, 1):
            #     new_task = threading.Thread(target=link_to_function, args=(f"{idx}.txt", obj))
            #     thread_task_list.append(new_task)
            #
            # for th in thread_task_list:
            #     th.start()
            #
            # for th in thread_task_list:
            #     th.join()

    class Databases:
        @staticmethod
        def read(db="postgres", password="31284bogdan", query="", autocommit=False):
            import psycopg2

            connection = None
            cursor = None
            records = None
            try:
                connection = psycopg2.connect(
                    user="postgres",
                    password=password,
                    host="127.0.0.1",  # 'localhost' \ '192.168.158.16'
                    port="5432",
                    dbname=db,
                )
                connection.autocommit = autocommit
                cursor = connection.cursor()
                cursor.execute(query)
                records = cursor.fetchall()
                connection.commit()
            except Exception as error:
                print(error)
                connection.rollback()
            finally:
                cursor.close()
                connection.close()

                return records
