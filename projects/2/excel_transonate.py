import openpyxl
# from openpyxl import load_workbook

# 1) читаю первую ячейку
# 2) читаю первую строку (+ for loop)
# 3) читаю вторую строку (+ for loop)
# 4) записываю в новый документ (с учётом перевёртывания)


workbook = openpyxl.load_workbook(filename='temp3/data1.xlsx')
# sheet_ranges = workbook['range names']
worksheet = workbook.active
print(worksheet['C2'], type(worksheet['C2']))  # <Cell 'Sheet1'.C2> <class 'openpyxl.cell.cell.Cell'>
print(worksheet['C2'].value)

print(worksheet.cell(2, 3))
print(worksheet.cell(2, 3).value)
# print(worksheet['D18'].value)

# start = 0
# stop = 10
# while start <= stop:

max_row = worksheet.max_row
max_column = worksheet.max_column

print("\n\n\n************\n\n\n")

# tuple1 = (12, 15, 17)
list1 = []
list2 = []

for number in range(1, 17+1, 1):  # 1 2 3 4 ... 10
    value1 = worksheet.cell(1, number).value
    list1.append(value1)
    value2 = worksheet[f"B{number}"].value
    list2.append(value2)
print(list1)
print(list2)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

index_row1 = 0
for row1 in list1:  # [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    index_row1 = index_row1 + 1
    new_worksheet.cell(index_row1, 1, row1)

index_row2 = 0
for row2 in list2:  # ['a', 'б', 'Bogdan', 'г', 'д', 'a', 'б', 'в', 'г', 'a', 'б', 'в', 'г', 'д', 'г', 'д', 'd']
    index_row2 = index_row2 + 1
    new_worksheet[f"B{index_row2}"] = row2

# for number in range(1, 17+1, 1):  # 1 2 3 4 ... 10
#     value1 = new_worksheet.cell(1, number).value
#     list1.append(value1)
#     value2 = new_worksheet.cell(2, number).value
#     list2.append(value2)

# new_workbook.save("temp3/new_data.xlsx")


################################################################

print("\n\n\n************\n\n\n\n\n\n\n\n\n")

workbook = openpyxl.load_workbook(filename='temp3/data1.xlsx')
worksheet = workbook.active

rows = []

for row in range(1, worksheet.max_row+1, 1):
    local_row = []
    for column in range(1, worksheet.max_column + 1, 1):
        value = worksheet.cell(row, column).value
        local_row.append(value)
    rows.append(local_row)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

for row_i, row in enumerate(rows, 1):
    for column_i, value in enumerate(row, 1):
        new_worksheet.cell(column_i, row_i, value)

new_workbook.save("temp3/new_data.xlsx")

