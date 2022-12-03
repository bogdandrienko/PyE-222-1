import datetime
import time
import json
import random
import math
import requests
import openpyxl
from openpyxl.utils import get_column_letter
import os

# print(datetime.datetime.now())  # 2022-12-03 10:38:24.112871
# time.sleep(5.5)  # задержка
# print(datetime.datetime.now())

# сериализация
# десериализация

dict1 = {
    "key": "value",
    "name": 'Bogdan',
    "age": 25,
    "teacher": True,
    "subjects": ["Python", "Django"]
}
print(dict1)
json_str = json.dumps(dict1)
print(json_str)
print(type(json_str))

# file = open("new.json", "w")
# file.write(json_str)
# file.close()

with open("new.json", "w", encoding='utf8') as file:
    file.write(json_str)
    # json.dump(dict1, file)


with open("new.txt", "w", encoding='utf8') as file:
    file.write("Привет")

with open("new.json", "rb") as file:
    # data = file.read()
    # print(data, type(data))
    dict2 = json.load(file)
    # dict2 = json.loads('''{"key": "value", "name": "Bogdan", "age": 25, "teacher": true, "subjects": ["Python", "Django"]}''')
    print(dict2, type(dict2))


print(random.randint(10, 100))
print(round(random.random() * 100, 2))  # 0.0 ... 1.0
print(math.floor(25.6))  # int
print(math.sqrt(25))  # 5.0
print(math.e)  # 2.718281828459045

# http - транспортный протокол
url = "https://jsonplaceholder.typicode.com/todos/1"
data = requests.get(url=url)
print(data, type(data))  # <Response [200]> <class 'requests.models.Response'>

print(data.content)
con = b'{\n  "userId": 1,\n  "id": 1,\n  "title": "delectus aut autem",\n  "completed": false\n}'

print(data.text, type(data.text))
text = '{"userId": 1,"id": 1,"title": "delectus aut autem","completed": false}'

print(data.json(), type(data.json()))  # {'userId': 1, 'id': 1, 'title': 'delectus aut autem', 'completed': False} <class 'dict'>
dict_data = data.json()

# with open("newnew.json", "wb") as file:
#     file.write(requests.get(url="https://jsonplaceholder.typicode.com/todos/66").content)

age = 20
message = "I'm " + str(age) + " age!"
message2 = f"I'm {age} age!"
print(message2)

# json_array = requests.get(url="https://jsonplaceholder.typicode.com/todos").json()
#
# for obj in json_array:
#     print(obj, type(obj))
#     with open(f'temp/new{obj["id"]}.json', "w") as file:
#         # file.write(json.dumps(obj))
#         json.dump(obj, file)

print("\n\n\n\n************\n\n\n\n")

# CRUD - create read update
workbook = openpyxl.load_workbook("new.xlsx")
worksheet = workbook.active

cell1 = worksheet["C2"]
print(cell1)  # <Cell 'Sheet1'.C2>
print(cell1.value)  # b_3

list5 = []
for number in range(1, 3+1):
    list5.append(worksheet[f"A{number}"].value)
print(list5)

matrix = []
rows = []
for row in range(1, worksheet.max_row + 1):
    local_rows = []
    for column in range(1, worksheet.max_column + 1):
        # value = worksheet[f"{get_column_letter(column)}{row}"].value
        local_rows.append(worksheet.cell(row, column).value)
    rows.append(local_rows)
print(rows)

vl = None
vl1 = ""
vl2 = False


new_workbook = openpyxl.Workbook()

new_worksheet = new_workbook.active

index = 0
for row in rows:
    for value in row:
        index += 1
        new_worksheet.cell(index, 1, value)

new_workbook.save("create.xlsx")
