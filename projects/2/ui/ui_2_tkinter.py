#####################
# сборка в exe

# pip install pyinstaller # обязательно внутри нужного виртуального окружения
# pyinstaller --onefile --windowed ui_1_tkinter.py
# pyinstaller --onedir --windowed ui_1_tkinter.py

#####################


# 1) ссылка, количество, название - программа для скачки картинок (Tkinter)
# 2) Парсер погоды (PySide2)
# 3) Парсер валюты (Pyqt6)

# from tkinter import *  # коллизии имён
from tkinter import ttk
import tkinter
import openpyxl

root = tkinter.Tk()
root.geometry("640x480")
frm = ttk.Frame(root, padding=10)

frm.grid()

ttk.Label(frm, text="Путь, где лежит файл: ").grid(column=0, row=0)
url = ttk.Entry(frm)
url.grid(column=1, row=0)
url.insert(0, "data/new.xlsx")

ttk.Label(frm, text="Колонки, для чтения: ").grid(column=0, row=1)
count = ttk.Entry(frm)
count.grid(column=1, row=1)
count.insert(0, "1, 4,7, 9")

ttk.Label(frm, text="Искомое: ").grid(column=0, row=2)
name = ttk.Entry(frm)
name.grid(column=1, row=2)


def start():
    print("start")

    url_for_download = str(url.get())
    # path = "data/new.xlsx"
    count_for_download = str(count.get()).split(sep=",")
    name_for_download = str(name.get())

    clear_count_for_download = []  # [1, 4, 7, 9]
    for i in count_for_download:  # ["1", " 4", "7 ", " 9"]
        int_value = int(i.strip())
        clear_count_for_download.append(int_value)

    # workbook = openpyxl.Workbook()  # TODO НОВЫЙ ФАЙЛ
    workbook = openpyxl.load_workbook(url_for_download)
    worksheet = workbook.active

    cell1 = worksheet.cell(1, 1).value
    print(cell1)

    rows1 = [  # rows
        [666, 2, 3],  # row [value, value, value]
        [66, 3, 4],
        [3, 4, 5],
    ]
    rows = []
    for row_index in range(1, 20 + 1):
        if row_index not in clear_count_for_download:  # 1, 3, 5, 9
            continue
        row = []
        for column_index in range(1, 10 + 1):
            cell = worksheet.cell(row_index, column_index)
            value = cell.value
            if value is None:
                value = 0
            # if value % 2 == 0:
            #     value = "Чётное"
            # else:
            #     value = "Нечётное"
            row.append(value)
        rows.append(row)
    print(rows)
    print("stop")


ttk.Label(frm, text="Количество совпадений: ").grid(column=2, row=3)
ttk.Button(frm, text="Старт", command=start).grid(column=1, row=3)

root.mainloop()

# путь и имя файла который нужно читать ("data/new.xlsx")
# перечисление колонок, которые нужно читать из файла ("1, 5,7, 9")
# элемент для поиска (+ вывод на экран количества совпадений)
# сортирует файл и перезаписывает в новый (по lambda)
