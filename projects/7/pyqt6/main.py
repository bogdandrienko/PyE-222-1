import datetime
import sys
import threading
import time
from PyQt6 import QtWidgets
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import FORMULAE


def timeit(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"Elapsed time: {round(end_time - start_time, 4)} seconds")
        return result
    return wrapper


class PyQtWindow(QtWidgets.QWidget):
    def __init__(self, window_title: str):
        super().__init__()

        global headers
        self.headers = headers

        self.play = True

        self.layout = QtWidgets.QGridLayout(self)

        self.label_url = QtWidgets.QLabel("url")
        self.layout.addWidget(self.label_url, 0, 0)

        self.line_edit_url = QtWidgets.QLineEdit("data.xlsx")
        self.layout.addWidget(self.line_edit_url, 1, 0)

        self.label_status = QtWidgets.QLabel("...")
        self.layout.addWidget(self.label_status, 0, 1)

        self.button_start = QtWidgets.QPushButton("запустить расчёт")
        self.button_start.clicked.connect(self.start)
        self.layout.addWidget(self.button_start, 1, 3)

        self.setWindowTitle(window_title)
        self.resize(640, 480)
        self.show()

    def start(self):
        self.play = True

        self.label_status.setText("идёт подсчёт")
        new_thread = threading.Thread(target=self.quest1)
        new_thread.start()

    def stop(self):
        self.play = False

    def finish(self, message="загрузка завершена"):
        self.label_status.setText(f"{message} [{datetime.datetime.now().strftime('%H:%M:%S')}]")

    @timeit
    def quest1(self):
        """
        1) вывести на экран сумму квадратов всех значений внутри excel-файла
        1.1 как прочитать данные с не активной рабочей книги в Python openpyxl
        1.2 как прочитать данные с первого столбца не активной рабочей книги в Python openpyxl
        1.3 вывести на экран сумму квадратов всех значений с первого столбца внутри excel-файла в Python openpyxl

        2) Написать декоратор, который будет замерять затраченное время в Python

        3) Запись результата в новый excel-файл в Python openpyxl

        4) Как форматировать значение с разделителями по пробелам, например "5000" как "5 000" в Python

        5) Как записать формулу на русском в Python openpyxl
        """
        try:
            filename: str = str(self.line_edit_url.text())

            workbook = openpyxl.load_workbook(f'data/{filename}', read_only=True)
            worksheet = workbook['данные']

            sum_of_squares: int = 0
            for row in worksheet.iter_rows(min_row=1, max_col=1, values_only=True):
                sum_of_squares += row[0] ** 2

            # sum_of_squares = str(sum_of_squares)
            # lenght_tar = len(str_sum_of_squares)
            # if 3 < lenght_tar < 6:
            #     str_sum_of_squares = str_sum_of_squares[3:] + " " + str_sum_of_squares[0:3]
            # elif 6 < lenght_tar < 9:
            #     str_sum_of_squares = str_sum_of_squares[6:] + " " + str_sum_of_squares[3:6] + " " + str_sum_of_squares[0:3]
            # elif 9 < lenght_tar < 12:
            #     str_sum_of_squares = str_sum_of_squares[9:] + " " + str_sum_of_squares[6:9] + " " + str_sum_of_squares[3:6] + " " + str_sum_of_squares[0:3]
            # print("НОВЫЙ ОТВЕТ: ", str_sum_of_squares)

            print("res", f'{sum_of_squares:,}')
            formatted_value = f'{sum_of_squares:,}'.replace(',', ' ')
            self.label_status.setText(f"{formatted_value}")

            new_workbook: Workbook = openpyxl.Workbook()
            new_worksheet: Worksheet = new_workbook.active

            new_worksheet.cell(row=1, column=1, value=sum_of_squares)
            new_worksheet.cell(row=1, column=2, value=formatted_value)
            new_worksheet.cell(row=2, column=1, value='=A1 + 1')

            new_workbook.save(f'data/new_{filename}')
        except Exception as error:
            print(f"error[{datetime.datetime.now()}]: ", error)


if __name__ == '__main__':
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/102.0.0.0 Safari/537.36'
    }

    pyqt_app = QtWidgets.QApplication([])
    pyqt_ui = PyQtWindow("пример приложений python")
    sys.exit(pyqt_app.exec())
