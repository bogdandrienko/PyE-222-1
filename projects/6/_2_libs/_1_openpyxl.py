########################################################################################################################
#
# # Нужно прочитать все(от 0 до 10) листы и записать на один вертикально
# import openpyxl
# from openpyxl.reader.excel import load_workbook
#
# # Import `load_workbook` module from `openpyxl`
# from openpyxl import load_workbook
#
# # Load in the workbook
# wb = load_workbook('data.xlsx')
# # sheet = wb.get_sheet_by_name('Лист1')
# #
# # for name in wb.sheetnames:
# #     print(name)
# sheet = wb['Лист1']
# a1 = sheet['A1'].value
# b1 = sheet['B1'].value
# c1 = sheet['C1'].value
# sheet = wb['Лист2']
# d1 = sheet['A1'].value
# e1 = sheet['B1'].value
# f1 = sheet['C1'].value
# sheet = wb['Лист3']
# g1 = sheet['A1'].value
# h1 = sheet['B1'].value
# j1 = sheet['C1'].value
# list1 = [[a1, b1, c1], [d1, e1, f1], [g1, h1, j1]]
# print(list1)
# # Нету динамичности(если имя рабочего листа или количества  то логика сломается)
# # Нету записи в новый файл
import datetime

########################################################################################################################

# todo Импорты библиотек и типов данных ################################################################################
import requests
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# todo Импорты библиотек и типов данных ################################################################################

#

# todo Получение "сырых" данных от api #################################################################################
url = "https://jsonplaceholder.typicode.com/posts"
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/102.0.0.0 Safari/537.36'
}
data = requests.get(url=url, headers=headers).json()
# print(data)


# todo Получение "сырых" данных от api #################################################################################

#

# todo Определение основной сущности ###################################################################################
class Post:
    def __init__(self, user_id: int, id: int, title: str, body: str):
        self.user_id = user_id
        self.id = id
        self.title = title
        self.body = body

    def __str__(self) -> str:
        return f"{self.title}({self.id})"

    def __repr__(self) -> str:
        return f"{self.title}({self.id})"

    def get_row(self) -> list[int | str]:
        """
        Конвертация сущности в массив для отображения в excel
        """
        return [self.user_id, self.id, self.title, self.body]


# todo Определение основной сущности ###################################################################################

#

# todo "Сериализация" данных с api в нашу сущность #####################################################################
# post1: Post = Post(user_id=1, id=1, title="title 1", body="body 1")
posts = []
for i in data:
    new_post = Post(user_id=i["userId"], id=i["id"], title=i["title"], body=i["body"])
    posts.append(new_post)
# print(posts)
# todo "Сериализация" данных с api в нашу сущность #####################################################################

#

# todo Создание новых рабочей книги и активация рабочего листа #########################################################
workbook = openpyxl.Workbook()  # создание нового файла в оперативной памяти
worksheet = workbook.active
# todo Создание новых рабочей книги и активация рабочего листа #########################################################

#

# todo Наполнение рабочего листа заголовками ###########################################################################
titles = ["#", "Id", "Title", "Body"]
column_index = 1
for value in titles:
    worksheet.cell(row=1, column=column_index, value=value)
    column_index += 1

# for column_index, value in enumerate(titles, 1):
#     worksheet.cell(row=1, column=column_index, value=value)
# todo Наполнение рабочего листа заголовками ###########################################################################

#

# todo Наполнение рабочего листа данными ###############################################################################
posts = [
    # голова туловище ноги ...
    ["#", "id", "title", "body"],  # студент 1

    # голова туловище ноги ...
    ["#", "id", "title", "body"],  # студент 2

    # голова туловище ноги ...
    ["#", "id", "title", "body"],  # студент 3

    # голова туловище ноги ...
    ["#", "id", "title", "body"],  # студент 4

    # голова туловище ноги ...
    ["#", "id", "title", "body"],  # студент 5
]

for row_index, post in enumerate(posts, 2):  # внешний цикл отвечает за переключение сущностей
    for column_index, value in enumerate(post.get_row(), 1):  # внутренний цикл отвечает за переключение атрибутов
        worksheet.cell(row=row_index, column=column_index, value=value)
# todo Наполнение рабочего листа данными ###############################################################################

#

# todo Сохранение и закрытие файла #####################################################################################
workbook.save(f"new_data_{datetime.datetime.now().strftime('%H-%M-%S')}.xlsx")
workbook.close()
# todo Сохранение и закрытие файла #####################################################################################
