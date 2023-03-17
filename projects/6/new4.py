from openpyxl import load_workbook, Workbook


file_name = 'data.xlsx'
workbook = load_workbook(file_name)
list1 = []
for sheetname in workbook.sheetnames:
    worksheet = workbook[sheetname]
    for j in range(1, worksheet.max_column+1):
        val = worksheet.cell(1, j).value
        if val is not None:
            list1.append(val)
print(list1)
new_wb = Workbook()
new_wb.save("new_data.xlsx")
